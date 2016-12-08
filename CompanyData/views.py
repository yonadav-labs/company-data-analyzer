# -*- coding: utf8 -*-

import pandas as pd
import numpy as np

import csv
import glob
import json
import os, sys
import datetime
import mimetypes

from django.utils.encoding import smart_str
from django.shortcuts import render
from django.conf import settings
from django.template.defaultfilters import slugify
from django.http import HttpResponseRedirect, HttpResponse
from django.http import JsonResponse
from django.forms.models import model_to_dict
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
from wsgiref.util import FileWrapper

from CompanyData.models import CompanyData


from fuzzywuzzy import process

from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl.utils.cell import (_get_column_letter)
from openpyxl.styles import Alignment


ORIGINAL_COLUMNS = ['Active Medstation', 'Alternate Items', 'Blocked Medstation', 
                    'CA Acct #', 'CA Doses Dispensed', 'Active Formulary Items', 
                    'CA Lines Refilled', 'CA Refill Qty', 'CA Scan Qty', 
                    'CA Scan Rate', 'CA Stock Outs', 'CA Vend to Refill Ratio', 
                    'CompanyID', 'Customer Name', 'DC', 'Ineligible Items', 
                    'MUS Formulary', 'Main Acct #', 'NON-CA Refill Qty', 
                    'NON-CA Scan Qty', 'Non-CA Doses Dispensed', 
                    'Non-CA Lines Refilled', 'Non-CA Scan Rate', 
                    'Non-CA Stock Outs', 'Non-CA Vend to Refill Ratio', 
                    'Refill Line Utilization', 'Report Date', 
                    'Total Doses Dispenses', 'Total Refill Qty', 'Total Refills', 
                    'Total Scan Qty', 'Total Stock Outs', 
                    'Total Stock Outs Per 100 Dispenses', 
                    'Total Vend Refill Ratio', 'Total Medstation', 
                    'Total Scan Rate', 'Utilization']


@login_required(login_url='/login/')
@csrf_exempt
def upload(request):
    if request.method == 'GET':
        return render(request, 'upload.html')
    else:
        myfile = request.FILES['files[]']
        # location = settings.BASE_DIR+"/media/test"
        # fs = FileSystemStorage(location=location)
        # filename = fs.save(myfile.name, myfile)

        result = {}
        result['name'] = myfile.name
        result['type'] = 'CSV'
        result['size'] = myfile.size
        # result['url'] = fs.url(filename)

        all_data = read_data(myfile)

        # save to database
        for item in all_data:
            # get rid of junk
            if item['companyid'] == 'nan {}'.format(item['ca_acct']):
                continue

            obj, created = CompanyData.objects.update_or_create(
                companyid=item['companyid'], report_date=item['report_date'],
                defaults=item,
            )

        return JsonResponse({'files': [result]}, safe=False)


def get_data(request):
    if request.method == 'GET':
        return render(request, 'export_data.html')
    else:
        company_id = request.POST.get('company_id').strip()
        company_data = CompanyData.objects.filter(companyid=company_id).order_by('-report_date')[:12]

        if not company_data:
            return HttpResponseRedirect('/get_data')        

        path = settings.BASE_DIR+"/media/company_data/"+datetime.datetime.now().strftime("{} (%Y-%m-%d %H:%M:%S).csv".format(company_id))
        result = open(path, 'w')

        result_csv_fields = model_to_dict(company_data[0], exclude=['id']).keys()
        result_csv = csv.DictWriter(result, fieldnames=result_csv_fields)
        result_csv.writeheader()

        for item in company_data:
            result_csv.writerow(model_to_dict(item, exclude=['id']))

        result.close()

        # wrapper = FileWrapper( open( path, "r" ) )
        # content_type = mimetypes.guess_type( path )[0]

        # response = HttpResponse(wrapper, content_type = content_type)
        # response['Content-Length'] = os.path.getsize( path ) # not FileField instance
        # response['Content-Disposition'] = 'attachment; filename=%s/' % smart_str( os.path.basename( path ) )
        return get_download_response(path)


@login_required(login_url='/login/')
def get_report(request):
    if request.method == 'GET':
        return render(request, 'export_report.html')
    else:
        company_id = request.POST.get('company_id')
        company_data = CompanyData.objects.filter(companyid=company_id).order_by('report_date')[:12]

        if not company_data:
            return HttpResponseRedirect('/get_report')

        path = get_report_(company_data)
        return get_download_response(path)


def get_download_response(path):
    wrapper = FileWrapper( open( path, "r" ) )
    content_type = mimetypes.guess_type( path )[0]

    response = HttpResponse(wrapper, content_type = content_type)
    response['Content-Length'] = os.path.getsize( path ) # not FileField instance
    response['Content-Disposition'] = 'attachment; filename=%s/' % smart_str( os.path.basename( path ) )
    return response


def get_report_(company_data):
    data = []
    for item in company_data:
        data.append(model_to_dict(item, exclude=['id']))

    QUERY_SET_KEY_TO_COLS = get_key_to_cols(data)

    df = pd.DataFrame(data)

    df.rename(index=str, columns=QUERY_SET_KEY_TO_COLS,inplace=True)

    # df.columns = ORIGINAL_COLUMNS        
    # df = df.drop('CompanyID', 1)

    # get file name
    # 2016-09 Monthly KPIs - Customer Name
    filename = settings.BASE_DIR+"/media/reports/"+"{} Monthly KPIs - {}.xlsx".format(list(df['Report Date'])[-1].strftime('%Y-%m'), df['Customer Name'][0])


    # format month names
    report_dates = [item.strftime('%b-%Y') for item in df['Report Date']]
    df = df.drop('Report Date', 1)
    dft = df.T
    dft.columns = report_dates
    # add empty YTD column
    dft['YTD'] = ['' for item in dft.index]
    # calc sum for YTD
    sum_rows = ['Alternate Items', 'CA Doses Dispensed',
                'Non-CA Doses Dispensed', 'Total Doses Dispenses',
                'Total Refills', 'CA Lines Refilled', 'Non-CA Lines Refilled',
                'Non-CA Stock Outs', 'CA Stock Outs', 'Total Stock Outs',
                'CA Scan Qty', 'CA Refill Qty', 'NON-CA Scan Qty',
                'NON-CA Refill Qty', 'Total Scan Qty', 'Total Refill Qty']
    for srow in sum_rows:                    
        try:
            dft.at[srow, 'YTD'] = dft.loc[srow].apply(pd.to_numeric).sum()
        except Exception, e:
            dft.at[srow, 'YTD'] = ''

    # calc division for YTD
    division_rows = [('CA Vend to Refill Ratio', 'CA Doses Dispensed', 'CA Lines Refilled'),
                     ('Non-CA Vend to Refill Ratio', 'Non-CA Doses Dispensed', 'Non-CA Lines Refilled'),
                     ('Total Vend Refill Ratio', 'Total Doses Dispenses', 'Total Refills'),
                     ('CA Scan Rate', 'CA Scan Qty', 'CA Refill Qty'),
                     ('Non-CA Scan Rate', 'NON-CA Scan Qty', 'NON-CA Refill Qty'),
                     ('Total Scan Rate', 'Total Scan Qty', 'Total Refill Qty')]

    for drow in division_rows:
        dft.at[drow[0], 'YTD'] = dft.at[drow[1], 'YTD'] / dft.at[drow[2], 'YTD']

    # create CY16
    cy16_index = ['Active Items on CardinalASSIST® Formulary', 
                  'CardinalASSIST® Formulary Utilization',
                  'Total Lines Auto Replenished',
                  'Percent Reduction in Lines Refilled by CardinalASSIST®',
                  'CardinalASSIST® Doses Dispensed',
                  'Non-CardinalASSIST® Doses Dispensed',
                  'Total ADM Doses (Vends) Dispensed',
                  'CardinalASSIST® Lines Refilled',
                  'Non-CardinalASSIST® Lines Refilled',
                  'Total ADM Lines Refilled',
                  'CardinalASSIST® Vend to Refill Ratio',
                  'Non-CardinalASSIST® Vend to Refill Ratio',
                  'Total Vend to Refill Ratio',
                  'CardinalASSIST® Stock Outs',
                  'Non-Cardinal-ASSIST® Stock Outs',
                  'Total ADM Stock Outs',
                  'Average CardinalASSIST® Stock Outs Per Day',
                  'Average Non-CardinalASSIST® Stock Outs Per Day',
                  'Total Average ADM Stock Outs Per Day',
                  'Stock Out Ratio (Per 100 ADM Dispenses)',
                  'CardinalASSIST® Items Scanned',
                  'CardinalASSIST® Items Refilled',
                  'CardinalASSIST® Scan Rate',
                  'Non-CardinalASSIST® Items Scanned',
                  'Non-CardinalASSIST® Items Refilled',
                  'Non-CardinalASSIST® Scan Rate',
                  'Total ADM Scan Rate']

    df_cy16 = pd.DataFrame(index=cy16_index, columns=report_dates+['YTD'])
    for month in report_dates:
        df_cy16.at['Active Items on CardinalASSIST® Formulary', month]  = dft.at['Active Formulary Items', month]
        df_cy16.at['CardinalASSIST® Formulary Utilization', month]  = dft.at['Utilization', month]
        df_cy16.at['Total Lines Auto Replenished', month]  = dft.at['CA Lines Refilled', month]
        df_cy16.at['Percent Reduction in Lines Refilled by CardinalASSIST®', month]  = (float(dft.at['Total Refills', month]) - float(dft.at['Non-CA Lines Refilled', month])) / float(dft.at['Total Refills', month])
        df_cy16.at['CardinalASSIST® Doses Dispensed', month]  = dft.at['CA Doses Dispensed', month]
        df_cy16.at['Non-CardinalASSIST® Doses Dispensed', month]  = dft.at['Non-CA Doses Dispensed', month]
        df_cy16.at['Total ADM Doses (Vends) Dispensed', month]  = float(dft.at['CA Doses Dispensed', month]) + float(dft.at['Non-CA Doses Dispensed', month])
        df_cy16.at['CardinalASSIST® Lines Refilled', month]  = dft.at['CA Lines Refilled', month]
        df_cy16.at['Non-CardinalASSIST® Lines Refilled', month]  = dft.at['Non-CA Lines Refilled', month]
        df_cy16.at['Total ADM Lines Refilled', month]  = float(dft.at['CA Lines Refilled', month]) + float(dft.at['Non-CA Lines Refilled', month])
        df_cy16.at['CardinalASSIST® Vend to Refill Ratio', month]  = dft.at['CA Vend to Refill Ratio', month]
        df_cy16.at['Non-CardinalASSIST® Vend to Refill Ratio', month]  = dft.at['Non-CA Vend to Refill Ratio', month]
        df_cy16.at['Total Vend to Refill Ratio', month]  = dft.at['Total Vend Refill Ratio', month]
        df_cy16.at['CardinalASSIST® Stock Outs', month]  = dft.at['CA Stock Outs', month]
        df_cy16.at['Non-Cardinal-ASSIST® Stock Outs', month]  = dft.at['Non-CA Stock Outs', month]
        df_cy16.at['Total ADM Stock Outs', month]  = dft.at['Total Stock Outs', month]
        df_cy16.at['Average CardinalASSIST® Stock Outs Per Day', month]  = float(dft.at['CA Stock Outs', month]) / 31
        df_cy16.at['Average Non-CardinalASSIST® Stock Outs Per Day', month]  = float(dft.at['Non-CA Stock Outs', month]) / 31
        df_cy16.at['Total Average ADM Stock Outs Per Day', month]  = float(dft.at['Total Stock Outs', month]) / 31
        df_cy16.at['Stock Out Ratio (Per 100 ADM Dispenses)', month]  = dft.at['Total Stock Outs Per 100 Dispenses', month]
        df_cy16.at['CardinalASSIST® Items Scanned', month]  = dft.at['CA Scan Qty', month]
        df_cy16.at['CardinalASSIST® Items Refilled', month]  = dft.at['CA Refill Qty', month]
        df_cy16.at['CardinalASSIST® Scan Rate', month]  = float(dft.at['CA Scan Qty', month]) / float(dft.at['CA Refill Qty', month])
        df_cy16.at['Non-CardinalASSIST® Items Scanned', month]  = dft.at['NON-CA Scan Qty', month]
        df_cy16.at['Non-CardinalASSIST® Items Refilled', month]  = dft.at['NON-CA Refill Qty', month]
        df_cy16.at['Non-CardinalASSIST® Scan Rate', month]  = float(dft.at['NON-CA Scan Qty', month]) / float(dft.at['NON-CA Refill Qty', month])
        df_cy16.at['Total ADM Scan Rate', month]  = (float(dft.at['CA Scan Qty', month]) + float(dft.at['NON-CA Scan Qty', month])) / (float(dft.at['CA Refill Qty', month])+float(dft.at['NON-CA Refill Qty', month]))

    df_cy16.at['Active Items on CardinalASSIST® Formulary', 'YTD']  = dft.at['Active Formulary Items', 'YTD']
    df_cy16.at['CardinalASSIST® Formulary Utilization', 'YTD']  = dft.at['Utilization', 'YTD']
    df_cy16.at['Total Lines Auto Replenished', 'YTD']  = dft.at['CA Lines Refilled', 'YTD']
    df_cy16.at['Percent Reduction in Lines Refilled by CardinalASSIST®', 'YTD']  = (float(dft.at['CA Lines Refilled', 'YTD']) - float(dft.at['Total Refills', 'YTD'])) / float(dft.at['CA Lines Refilled', 'YTD'])
    df_cy16.at['CardinalASSIST® Doses Dispensed', 'YTD']  = dft.at['CA Doses Dispensed', 'YTD']
    df_cy16.at['Non-CardinalASSIST® Doses Dispensed', 'YTD']  = dft.at['Non-CA Doses Dispensed', 'YTD']
    df_cy16.at['Total ADM Doses (Vends) Dispensed', 'YTD']  = dft.at['Total Doses Dispenses', 'YTD']
    df_cy16.at['CardinalASSIST® Lines Refilled', 'YTD']  = dft.at['CA Lines Refilled', 'YTD']
    df_cy16.at['Non-CardinalASSIST® Lines Refilled', 'YTD']  = dft.at['Non-CA Lines Refilled', 'YTD']
    df_cy16.at['Total ADM Lines Refilled', 'YTD']  = dft.at['Total Refills', 'YTD']
    df_cy16.at['CardinalASSIST® Vend to Refill Ratio', 'YTD']  = dft.at['CA Vend to Refill Ratio', 'YTD']
    df_cy16.at['Non-CardinalASSIST® Vend to Refill Ratio', 'YTD']  = dft.at['Non-CA Vend to Refill Ratio', 'YTD']
    df_cy16.at['Total Vend to Refill Ratio', 'YTD']  = dft.at['Total Vend Refill Ratio', 'YTD']
    df_cy16.at['CardinalASSIST® Stock Outs', 'YTD']  = dft.at['CA Stock Outs', 'YTD']
    df_cy16.at['Non-Cardinal-ASSIST® Stock Outs', 'YTD']  = dft.at['Non-CA Stock Outs', 'YTD']
    df_cy16.at['Total ADM Stock Outs', 'YTD']  = dft.at['Total Stock Outs', 'YTD']
    df_cy16.at['Average CardinalASSIST® Stock Outs Per Day', 'YTD']  = float(dft.at['CA Stock Outs', 'YTD']) / 31
    df_cy16.at['Average Non-CardinalASSIST® Stock Outs Per Day', 'YTD']  = float(dft.at['Non-CA Stock Outs', 'YTD']) / 31
    df_cy16.at['Total Average ADM Stock Outs Per Day', 'YTD']  = float(dft.at['Total Stock Outs', 'YTD']) / 31
    df_cy16.at['Stock Out Ratio (Per 100 ADM Dispenses)', 'YTD']  = dft.at['Total Stock Outs Per 100 Dispenses', 'YTD']
    df_cy16.at['CardinalASSIST® Items Scanned', 'YTD']  = dft.at['CA Scan Qty', 'YTD']
    df_cy16.at['CardinalASSIST® Items Refilled', 'YTD']  = dft.at['CA Refill Qty', 'YTD']
    df_cy16.at['CardinalASSIST® Scan Rate', 'YTD']  = dft.at['CA Scan Rate', 'YTD']
    df_cy16.at['Non-CardinalASSIST® Items Scanned', 'YTD']  = dft.at['NON-CA Scan Qty', 'YTD']
    df_cy16.at['Non-CardinalASSIST® Items Refilled', 'YTD']  = dft.at['NON-CA Refill Qty', 'YTD']
    df_cy16.at['Non-CardinalASSIST® Scan Rate', 'YTD']  = dft.at['Non-CA Scan Rate', 'YTD']
    df_cy16.at['Total ADM Scan Rate', 'YTD']  = dft.at['Total Scan Rate', 'YTD']

    
    # Generate Report
    report_from_template(df_cy16, filename)

    return filename


def read_data(csv_file):
    """
    Read monthly company data from csv file and returns list of dictionary
    for storing into database
    """
    # Importing ALL Data and creating a Master DataFrame
    all_data = pd.DataFrame()
    # for f in glob.glob(settings.BASE_DIR+"/media/*MonthlyStatistics.*"):
        # date_str = f.split("/")[-1:][0][:7]
    date_str = csv_file.name[:7]
    df = pd.read_csv(csv_file)
    
    # Create the Datetime Stamp from the filename
    df['Report Date'] = date_str+'-01'
    # df['Report Date'] = str(pd.to_datetime(df['Report Date']))
    
    # Create the Company ID
    company_id_list = []
    # for i in zip(list(df['Customer Name']), list(df['CA Acct #']),list(df['Main Acct #'])):    
    for i in zip(list(df['Customer Name']), list(df['CA Acct #'])):
        tmp_id = "{} {}".format(i[0], i[1])
        company_id_list.append(tmp_id)
    
    # Company ID
    df['CompanyID'] = company_id_list
    
    #Return the master dataframe
    all_data = all_data.append(df,ignore_index=True)
        
    # Delete Unknown Field Name
    # Think this must be an import/ df.append issue (maybe you can figure it out)
    try:
        del all_data['Unnamed: 0']
    except Exception, e:
        pass
    
    new_cols = []
    for item in all_data.columns:
        new_cols.append(slugify(item).replace('-', '_'))
    all_data.columns = new_cols

    # print json.dumps(all_data.T.to_dict().values(), indent=4)
    return all_data.T.to_dict().values()
    # return all_data


def user_login(request):
    message = ''

    if request.method == 'POST':
        next_url = request.GET.get('next', '/')
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)

        if user is not None:
            login(request, user)
            return HttpResponseRedirect(next_url)
        else:
            message = 'Your login credential is not correct! Please try again.'
            
    return render(request, 'login.html', {
        'message': message,
        'l_block': 'login'
    })


def user_signup(request):
    message = ''

    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']

        try:
            User.objects.create_user(username, email, password)
            user = authenticate(username=username, password=password)
            login(request, user)
            return HttpResponseRedirect('/')
        except Exception, e:
            print e
            message = 'Your username is already used. Please try with another one!'
            
    return render(request, 'login.html', {
        'message': message,
        'l_block': 'signup'
    })


def user_logout(request):
    logout(request)
    return HttpResponseRedirect('/login')


def get_all_cells_A1_reference_style(dimension):
    start_column = re.sub(r"\d","",dimension.split(":")[0])
    end_column = re.sub(r"\d","",dimension.split(":")[1])
    
    start_row = int(re.sub(r"[A-Z]","",dimension.split(":")[0]))
    end_row = int(re.sub(r"[A-Z]","",dimension.split(":")[1]))

    # Create Range of Columns
    range_columns = []
    for i in range(ord(start_column), ord(end_column)+1):
        range_columns.append(chr(i))


    # Create Range of Rows
    range_rows = list(range(start_row,end_row+1))

    # Create Master List Combining Range of Columns and Rows (should have created this using a list comprehension)
    list_all_cells_in_worksheet = []
    for c in range_columns:
        for r in range_rows:
            list_all_cells_in_worksheet.append(str(c)+str(r))

    return list_all_cells_in_worksheet


def report_from_template(df_cy16, filename):
    # Load the Template
    wb = load_workbook(settings.BASE_DIR+"/media/reports/KPI_Template.xlsx")

    # select the sheet
    ws = wb.get_sheet_by_name("CY16")

    # Get the last row
    end_row = ws.max_row

    #####################################################################################################
    #
    # Integrate Code Above
    #
    #####################################################################################################

    index_col_cells, col_rows = get_cell_refs_for_index_and_cols()

    #####################################################################################################

    # Styles
    # We need to make sure to manually edit the template file that we are using in order to ensure that 
    # our formatting during the script works appropriately

    # Insert Column Headers
    for i, col in enumerate(list(df_cy16.columns)):
        ws[str(_get_column_letter(i+2)+"1")].value = col

    # Index of DataFrame
    zipped_ref_and_index_values = zip(index_col_cells,list(df_cy16.index))

    for input_ref_value in zipped_ref_and_index_values: 
        ws[input_ref_value[0]].value = input_ref_value[1]
        ws[input_ref_value[0]].alignment = Alignment(horizontal='right') 

    # All Columns in DataFrame
    for i, col_name in enumerate(list(df_cy16.columns)):
        col_cell_refs = [str(_get_column_letter(i+2)+row_ref) for row_ref in col_rows]
        zipped_ref_and_col_values = zip(col_cell_refs,list(df_cy16.ix[:,i].values))

        for input_ref_value in zipped_ref_and_col_values: 
            ws[input_ref_value[0]].value = input_ref_value[1]
            ws[input_ref_value[0]].alignment = Alignment(horizontal='right') 

    # Insert Remaining Table Headers specific to this Table
    table_headers = ['Program Summary','ADM Optimization','ADM Stock Outs','ADM Scan Rate on Refill Transactions']
    table_headers_refs = ['A2','A8','A21','A32']
    zipped_ref_and_table_headers = zip(table_headers_refs,table_headers)

    for input_ref_value in zipped_ref_and_table_headers: 
        ws[input_ref_value[0]].value = input_ref_value[1]
        ws[input_ref_value[0]].alignment = Alignment(horizontal='right')

    wb.save(filename)


def get_cell_refs_for_index_and_cols():

    # Load the Original workbook
    wb = load_workbook(settings.BASE_DIR+"/media/reports/KPI_Template.xlsx")

    # select the sheet
    ws = wb.get_sheet_by_name("CY16")

    # Get the last row
    end_row = ws.max_row

    # Get Index Mapping
    index_col_cells_all = [str("A"+str(i)) for i in range(1,end_row+1)]

    table_headers_or_None_index = [None,'Program Summary','ADM Optimization','ADM Stock Outs','ADM Scan Rate on Refill Transactions','=Data!B5']

    index_col_cells = []
    for cell in index_col_cells_all:
        if ws[cell].value not in table_headers_or_None_index: 
            index_col_cells.append(cell)

    #Cells for which we need Index values inserted
    # index_col_cells


    # Get Column Mapping
    col_cells_all = [str("D"+str(i)) for i in range(1,end_row+1)]

    table_headers_or_None_cols = [None,'=Data!D1']

    col_cells = []
    for cell in col_cells_all:
        if ws[cell].value not in table_headers_or_None_cols: 
            col_cells.append(cell)

    #Cell Rows for which we need Column values inserted
    col_rows = [ref[1:] for ref in col_cells]

    return index_col_cells, col_rows


def get_key_to_cols(query_set):
    query_set_keys = query_set[0].keys()
    cols = []
    for qsk in query_set_keys:
        cols.append(process.extractOne(qsk,choices=ORIGINAL_COLUMNS,score_cutoff=70)[0])

    # Create a dictionary mapping of these two lists
    QUERY_SET_KEY_TO_COLS = dict(zip(query_set_keys, cols))

    # Remove "companyid" from dictionary
    QUERY_SET_KEY_TO_COLS.pop('companyid')

    # Fix an error from the fuzzywuzzy matching in the dictionary
    QUERY_SET_KEY_TO_COLS['ca_formulary'] = 'Active Formulary Items'
    return QUERY_SET_KEY_TO_COLS

